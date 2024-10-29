from django.shortcuts import render, redirect
from .models import Articles
from .forms import ArticlesForms
from django.views.generic import DetailView, UpdateView, DeleteView
from django.conf import settings
import os
from django.http import FileResponse
from django.core.files.storage import FileSystemStorage

#Для таймера
from django.utils import timezone
from datetime import timedelta

import time
import threading

#Для таблицы
from .tables import ArticlesTable
from django_tables2 import RequestConfig

#Для логирования
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import PermissionRequiredMixin

#exsel
import openpyxl


@login_required
def news_home(request):
   #delay = 5  # время между вызовами функции в секундах, в данном примере - 5 c
   # while True:
   #time.sleep(delay)
   #redirect(request)
   # вызываемая в отдельном потоке функция в ней и производим действия из следующего шага
   # thread = threading.Thread(target=news_home, args=(request,))
   # thread.start()
   #print(request)

   # Проверяем есть ли у данного пользователя разрешение для просмотра
   # Если такого разрешения нет, то выкидываем исключение PermissionDenied
    if not request.user.has_perm('news.view_articles'):
       raise PermissionDenied

    news = Articles.objects.all()
    #news = Articles.objects.order_by('title')
    Time=timezone.now()+timedelta(hours=5)
    #Time=Time+timedelta(seconds=60)
    Time=Time.strftime("%a %d-%m-%Y %H:%M:%S")
    data={
        'news': news,
        'Time': Time,
    }
    return render(request, 'news/news_home.html', data)


def news_home_tables(request):
    table = ArticlesTable(Articles.objects.all())
    RequestConfig(request).configure(table)

    data={
        'table': table,
    }
    return render(request, 'news/news_home_tables.html', data)


class NewDetailView(PermissionRequiredMixin, DetailView):
#только тому пользователю у которого есть просмотр доступ
 permission_required = 'news.change_articles'
 model=Articles
 template_name = 'news/detail_views.html'
 context_object_name = 'new_article'




class NewUpdateView(PermissionRequiredMixin, UpdateView):
    # только тому пользователю у которого есть просмотр доступ
    permission_required = 'news.change_articles'
    model = Articles
    template_name = 'news/create.html'
    #context_object_name = 'UpdateView'
    #fields = ['title','anons','full_text','date']
    form_class = ArticlesForms



class NewDeleteView(DeleteView):
    model = Articles
    success_url = '/news/'
    template_name = 'news/news-delete.html'


@login_required
#только тому пользователю у которого есть добавление доступ
@permission_required('news.add_articles', raise_exception=True)
def create(request):
    error=''
    if request.method == 'POST':
        form = ArticlesForms(request.POST)
        if form.is_valid():
            #title = form.cleaned_data["title"]
            #print(title)
            form.save()
            return redirect('home')
        else: error='Ошибка формы'


    form= ArticlesForms()
    data={
        'form': form,
        'error': error,
    }
    return render(request, 'news/create.html', data)

  # Таймер
def do_news_taimer(request):
    delay = 3  # время между вызовами функции в секундах, в данном примере - 1 c
    #print(request)
    while True:
        time.sleep(delay)
        redirect('news_home')
        print("Отправил 4")
        #news_home('/news/')
        # вызываемая в отдельном потоке функция в ней и производим действия из следующего шага
        #thread = threading.Thread(target=do_news_taimer, args=("/",))
        #thread.start()

    pass

def read_save_EXSEL(request):
    #Чтение и запись Ексель
    file_new_path = os.path.join(settings.MEDIA_ROOT,"", "Temp3.xlsx")
    print(file_new_path)
    if request.method == 'POST':
        print("--POST--")
        if request.FILES:
            # получаем загруженный файл
            file = request.FILES['file']
            print('Name:' + file.name)
            fs = FileSystemStorage()
            # сохраняем на файловой системе
            filename = fs.save('Temp3.xlsx', file)
            #print('filename:' + filename)
            #file_new_path=fs.url(filename)
            #print(file_new_path)
            wb = openpyxl.load_workbook(file_new_path)
            print(wb.sheetnames)
            sh1 = wb['Лист1']
            row = sh1.max_row
            column = sh1.max_column
            # Чтение
            for i in range(2, row + 1):
            # print(i.value)
               print(f"=-=-=-Чтение строки -=-={i}")
               print(sh1.cell(row=i, column=1).value)
               print(sh1.cell(row=i, column=2).value)
               print(sh1.cell(row=i, column=3).value)
              # запись
            for i in range(2, row + 1):
               print(f"=-=-=-запись строки -=-={i}")
               sh1.cell(row=i, column=1).value = "1"

            wb.save(file_new_path)

    else:
            error='Ошибка формы'

    return redirect('news_home')



def news_save_EXSEL(request):
    file_path = os.path.join(settings.MEDIA_ROOT,"", "Temp2.xlsx")
    # создаем книгу
    wb = openpyxl.Workbook()
    # делаем единственный лист активным
    #ws = wb.active
    # вставить рабочий лист в конец (по умолчанию)
    ws1 = wb.create_sheet("Mysheet")
    # переименуем лист
    ws1.title = "NewPage"
    #Рабочий лист можно получить, используя его имя в качестве ключа экземпляра созданной книги Exce
    sh1 = wb["NewPage"]
    # удаление листов книги
    #wb.remove(wb['NewPage'])

    #запись
    sh1.cell(row=1, column=1).value="КОД"
    sh1.cell(row=1, column=2).value = "Артикул"
    sh1.cell(row=1, column=3).value = "Сумма"

    wb.save(file_path)

    return FileResponse(open(file_path,'rb'))

